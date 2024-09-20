import openpyxl
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

try:
    planilha_result = openpyxl.load_workbook('planilha fechamento.xlsx')
except FileNotFoundError:
    planilha_result = openpyxl.Workbook()
    pagina_resultado = planilha_result.active
    pagina_resultado.title = 'Resultado da planilha'
    pagina_resultado.append(['Nome', 'CPF', 'Status', 'Vencimento', 'Método de Pagamento'])
else:
    pagina_resultado = planilha_result['Sheet1']

driver = webdriver.Edge()
driver.get('https://consultcpf-devaprender.netlify.app')


for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    sleep(5)
    campo_pesquica = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquica.clear()
    campo_pesquica.send_keys(cpf)
    sleep(1)
    campo_consulta = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    campo_consulta.click()
    sleep(4)
    status_elemento = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    status = status_elemento.text

    print(f"Status capturado: '{status}'")
   
    if status == 'em dia':
        vencimento_elemento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_de_pagamento_elemento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
        data_pagamento = vencimento_elemento.text.split()[3]
        metodo_pagamento = metodo_de_pagamento_elemento.text.split()[3]
    else:
        data_pagamento = "N/A"
        metodo_pagamento = "N/A"
    
    pagina_resultado.append([nome, cpf, status, data_pagamento, metodo_pagamento])
    planilha_result.save('planilha fechamento.xlsx')
